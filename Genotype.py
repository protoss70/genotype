import openpyxl
from os import getcwd as getPath
from os import rename


#when FE is written in the document it means the fonction or the variable is only used in front end and not necesary for
#this code

#FE
errors = []
#Language indicator n
n = 1
#if dataPrivacy = 0 then names arent hidden if 1 names are hidden inside system inside data1 as 8 bit seperated by ","
dataPrivacy = 0

#Lang[target sentence][n]
#n = Language in use (0 for Turkish and 1 for English)
Lang = [
    ["Excel üzerinden hatalı giriş!\nisim: {}\nid: {}", "Invalid input!\nname: {}\nid: {}"],
    ["{} id'li ve {} isimli kişinin\nanne veya babasının id'si yanlış girilmiş",
     "Person with id: {} and name {}'s\nFathers or mothers id is invalid"],
    ["Rapor", "Report"],#2
    ["İsim: {}\nİd: {}\nCinsiyet: {}\nAnne adı: {}\t\tAnne id'si: {}\nBaba adı: {}\t\tBaba id'si: {}\nFenotip: {}\nGenotip:\n",
     "Name: {}\nId: {}\nSex: {}\nMothers name: {}\t\tMothers id: {}\nFathers name: {}\t\tFathers id: {}\nPhenotype: {}\nGenotype:\n"],
    ["İsim: {}\nİd: {}\nCinsiyet: {}\nFenotip: {}\nGenotip:\n",
     "Name: {}\nId: {}\nSex: {}\nPhenotype: {}\nGenotype:\n"],#4
    ["Sistem", "System"],#5
    ["Raporlar", "Reports"],#6
    ["Hastalık Raporu", "Disease Report"],#7
    ["{} ile {}, çocuk raporu", "{} and {}, child report"],#8
    ["Erkek çocuk", "Male child"],#9
    ["Kız çocuk", "Female child"],#10
    ["Lütfen Excel'i kapatıp tekrar deneyin", "Please Close the Excel and try again"]
]
#-------------------------------Class'-------------------------------


class Patient:
    #all instances of Patient class is kept inside Patients
    Patients = []
    #All mutated patients are kept inside MutatedPatients and they are filled inside main_func()
    MutatedPatients = []
    #BigG is short for biggest generation and finds the youngest person with the longest family tree
    BigG = 1
    '''
    Father and Mother are ids
    '''
    def __init__(self, Fullname, Father, Mother, GenoType, PhenoType, id, Male=False):
        self.Fullname = Fullname
        self.Father = Father
        self.Mother = Mother
        self.GenoType = GenoType
        self.PhenoType = PhenoType
        self.Children = []
        self.id = id
        self.Male = Male
        self.Generation = None
        if Mother != None and Father != None:
            self.Mother.Children.append(self)
            self.Father.Children.append(self)
        Patient.Patients.append(self)

    def gene_merger(self, new_Genes):
        '''
        Merges self's possible genotypes with new_Genes
        input: self.GenoType = ["Ao", "AA"]
               new_Genes     =  [['RR'], ['Rr']]

        result: self.GenoType = ['AARR', 'AARr', 'AoRR', 'AoRr']
        '''
        if self.GenoType == None:
            self.GenoType = new_Genes
        else:
            pos = []
            for i in self.GenoType:
                newString = ""
                for a in i:
                    newString += a
                for b in new_Genes:
                    new1 = ""
                    for c in b:
                        new1 += c
                    pos.append(newString+new1)
            self.GenoType = pos

    @staticmethod
    def Crosser(Genos, toMerge, gonosomal=False, Male=False):
        '''
        Crosses genes of 2 people
        input: Genos = []
                toMerge = [[['A', 'o'], ['A', 'o'], ['A', 'o'],  ['A', 'o']],
                            [['o', 'o'], ['o', 'o']]]
                index 0 = Mothers genes for gene type
                index 1 = Fathers genes for gene type
                gonomosal for gene type and male for sex

        output: newGeno = [['Ao', 16], ['oo', 16]]
                without the Counter method newGeno = ['Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao',
                                                'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo',
                                                'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo', 'Ao', 'oo']

                After Counter: newGeno = [['Ao', 16], ['oo', 16]]
                first index ["Ao", 16] shows the gene and the second index shows how many times
                that combination of genes were created

                in the second run the second indexs sum

                Second run: newGeno = [['Aorr', 512], ['oorr', 512]]
        '''
        newGeno = []
        a = toMerge

        if len(Genos) == 0 and gonosomal == False:
            for i in a[0]:
                for b in a[1]:
                    newGeno.append(i[0] + b[0])
                    newGeno.append(i[1] + b[0])
                    newGeno.append(i[0] + b[1])
                    newGeno.append(i[1] + b[1])
            newGeno = Counter(newGeno, True)
            return newGeno

        elif gonosomal == False:
            for i in Genos:
                for b in a[0]:

                    for c in a[1]:
                        newGeno.append([i[0] + b[0] + c[0], i[1]])
                        newGeno.append([i[0] + b[1] + c[0], i[1]])
                        newGeno.append([i[0] + b[0] + c[1], i[1]])
                        newGeno.append([i[0] + b[1] + c[1], i[1]])

            newGeno = Counter(newGeno, mode=True)

            return newGeno
        elif gonosomal and len(Genos) == 0:
            if Male:
                if len(a[0]) == 0:
                    for b in a[1]:
                        newGeno.append(b[0])
                    return newGeno
                else:
                    for b in a[0]:
                        newGeno.append(b[0])
                        newGeno.append(b[1])
                    return newGeno
            else:
                for b in a[0]:
                    for c in a[1]:
                        newGeno.append(b[0] + c[0])
                        newGeno.append(b[1] + c[0])

                newGeno = Counter(newGeno, True)
                return newGeno

        elif gonosomal and len(Genos) != 0:
            if Male:
                if len(a[0]) == 0:
                    for i in Genos:
                        for b in a[1]:
                            newGeno.append([i[0] + b[0], i[1]])
                else:
                    for i in Genos:
                        for b in a[0]:
                            for c in a[1]:
                                newGeno.append([i[0] + b[0], i[1]])
                                newGeno.append([i[0] + b[1], i[1]])
                                newGeno.append([i[0] + c[0], i[1]])
            else:
                if len(a[0]) != 0:
                    for i in Genos:
                        for b in a[0]:
                            for c in a[1]:
                                newGeno.append([i[0] + b[0] + c[0], i[1]])
                                newGeno.append([i[0] + b[1] + c[0], i[1]])
            newGeno = Counter(newGeno, mode=True)
            return newGeno


class Disease:
    '''
    Diseases keeps all the instances of Disease class
    '''
    Diseases = []


    def __init__(self, name, Phenotype):
        self.name = name
        self.Phenotype = Phenotype

        Disease.Diseases.append(self)

    def disease_posibility(self, GenoType, Male):
        '''
        Takes a Genotype as a string and returns True if
        person has the disease otherwise returns False

        Male is the sex of the Genotype owner

        Compares Genotype with self.Phenotype and finds the result
        '''
        for gene in Gene.Genes:
            dis = gene.gene_in_pheno(self.Phenotype)
            ph = gene.gene_in_genos([GenoType], Male)

            if len(ph) > 0:
                ph = gene.geno_to_pheno(ph[0])
            if len(dis) > 0:
                dis = gene.geno_to_pheno(dis)
            if len(ph) > 0 and len(dis) > 0:
                if ph != dis:
                    return False
            if gene.gonosomal == True and len(dis) != 0 and len(ph) == 0:
                return False
        return True


class Gene:
    #Genes keeps all the instances of Gene Class
    Genes = []
    '''
    Dominance is the dominance of the letters for example letters = ["A", "B", "o"] and dominance = [1,1,0]
    this means A:1 B:1 and o:0 so if the Genotype is "Ao" the Phenotype is "A"
    but if Genotype is AB then Phenotype is "AB"
    '''
    def __init__(self, name, dominance, letters, gonosomal=False):
        self.name = name
        self.dominance = dominance
        self.letters = letters
        self.gonosomal = gonosomal
        Gene.Genes.append(self)

    def geno_to_pheno(self, geno):
        '''
        Takes relevant genes

        if self.letters = ["A", "B", "o"]
            self.dominance = [1,1,0]
        input: geno = ["A", "o"]

        output: geno = "A"
        '''
        if len(geno) > 1:
            numerical = ""
            for i in range(len(geno)):
                for b in range(len(self.letters)):
                    if geno[i] == self.letters[b]:
                        numerical += str(self.dominance[b])
            if int(numerical[0]) > int(numerical[1]):
                return geno[0]
            elif int(numerical[0]) == int(numerical[1]):
                if geno[0] == geno[1]:
                    return geno[0]
                else:
                    return geno[0] + geno[1]
            elif int(numerical[0]) < int(numerical[1]):
                return geno[1]
        else:
            return geno[0]

    def pheno_to_geno(self, gene, male=False):
        '''
        Takes a Phenotype as a list

        input: gene = ["R"]

        output: [['RR'], ['Rr']]


        input: gene = ["A", "B"]

        output: ["AB"]
        '''

        if male == False:
            posibilities = []
            point = None
            if len(gene) > 1:
                return [["{}{}".format(gene[0], gene[1])]]
            else:
                for i in gene:
                    for a in range(len(self.letters)):
                        if i == self.letters[a]:
                            point = a
                posibilities.append(["{}{}".format(self.letters[point], self.letters[point])])
                for i in range(len(self.dominance)):
                    if self.dominance[point] > self.dominance[i]:
                        posibilities.append(["{}{}".format(self.letters[point], self.letters[i])])

            return posibilities
        else:
            return [["{}".format(gene[0])]]

    def gene_in_genos(self, genes, Male=False):
        '''
        Takes a list of Genotypes

        self determines which genes will return from Genotypes

        input: genes = ['ooRRmmXHXHXrXr', 'ooRRmmXHXhXrXr', 'ooRrmmXHXHXrXr', 'ooRrmmXHXhXrXr']

        output: [['R', 'R'], ['R', 'R'], ['R', 'r'], ['R', 'r']]
        '''
        l = []
        gene = None
        for i in genes:
            first = []
            b = 0
            x = b
            while b < len(i):
                if i[b] == "Y" or i[b] == "X":
                    x = b + 2
                else:
                    x = b + 1
                try:
                    while True:

                        int(i[x])
                        x += 1
                except:
                    gene = Gene.gene_finder(i[b:x])
                    if gene == self:
                        if self.gonosomal == False:
                            if len(first) == 0:
                                first.append(i[b:x])
                                b = x
                            else:
                                first.append(i[b:x])
                                l.append(first)
                                b = len(i)
                        elif Male == True:
                            l.append([i[b:x]])
                            b = len(i)
                        else:
                            if len(first) == 0:
                                first.append(i[b:x])
                                b = x
                            else:
                                first.append(i[b:x])
                                l.append(first)
                                b = len(i)
                    else:
                        b = x
        return l

    def gene_in_pheno(self, pheno):
        '''
        Takes a Phenotype as a string
        Looks for self.letters in pheno
        if cannot find it returns []

        self.letters = ["YO", "Yo"]
        input: pheno = XRXhYo
        output: ['Yo']
        '''
        l = []
        gene = None
        b = 0
        x = b
        while b < len(pheno):
            if pheno[b] == "Y" or pheno[b] == "X":
                x = b + 2
            else:
                x = b + 1
            try:
                while True:

                    int(pheno[x])
                    x += 1
            except:
                gene = Gene.gene_finder(pheno[b:x])
                if gene == self:
                    l.append(pheno[b:x])
                b = x
        return l

    @classmethod
    def gene_finder(cls, letter):
        '''
        looks trough all genes and finds the one with input letters

        input: letter = "Xh"
        output = Gene Class (gene.name = "Hemophili")
        '''
        genes = Gene.Genes
        for gene in genes:
            for i in gene.letters:
                if i == letter:
                    return gene
        return None


#--------------------------funcs--------------------------------------


def Counter(list, Count=False, mode=False):
    if Count is False and mode == False:
        '''
        removes the same elements from the list
        input: ["AB", "Ao", "AB", "Ao","AB", "AA"]
        
        output: ["AB", "Ao", "AA"] 
        '''
        a = []
        for i in list:
            g = False
            for b in a:
                if i == b:
                    g = True
            if g is False:
                a.append(i)
        return a
    elif Count and mode==False:
        '''
        Counts all list elements
        input: ["AB", "Ao", "AB", "Ao","AB", "AA"]
        
        output: [["AB", 3], ["Ao", 2], ["AA", 1]]
        '''
        a = []
        for i in list:
            g = False
            for b in a:
                if i == b[0]:
                    b[1] += 1
                    g = True
            if g is False:
                a.append([i, 1])
        return a
    elif mode == True and Count == False:
        #["ABRr", 200], ["ABRr", 200], ["ABrr", 200]
        '''
        Given a list with elements first index is a string and the second index is the value
        if the first index is repeated then sums the second index with the repeated elements secon index
        input: ["ABRr", 200], ["ABRr", 200], ["ABrr", 200]
        
        output: [["ABRr", 400], ["ABrr", 200]]
        '''
        a = []
        for i in list:
            g = False
            for b in a:
                if i[0] == b[0]:
                    b[1] += i[1]
                    g = True
            if g == False:
                a.append([i[0], i[1]])
        return a


def print_values():
    #this function is used to visualize the data
    print("\n")
    for person in Patient.Patients:
        print("{}\tGenoType: {}".format(person.Fullname, person.GenoType))

    print("\nMutated Patients")
    for person in Patient.MutatedPatients:
        print(person.Fullname)


def disease_posibility_person(id):
    '''
    input: id of a person

    output: if id is wrong False
            else returns a list of disease names and percentage of having the disease
    '''
    target = None
    for person in Patient.Patients:
        if person.id == id:
            target = person
    if target == None:
        return False
    else:
        Genos = target.GenoType
        Male = target.Male
        all = []
        for disease in Disease.Diseases:
            a = 0
            b = 0
            for gene in Genos:
                if disease.disease_posibility(gene, Male):
                    a += 1
                    b += 1
                else:
                    b += 1
            result = round((a / b) * 100, 2)
            all.append([disease.name, result])
        return all


def child_disease_posibility(father_id, mother_id):
    '''
    Takes an id of father and a mother
    if mother or father's id is wrong returns False
    else Crosses the genes via Patient.Crosser and produces both male and female childs

    example male child: [['AorrXHXRYoMM', 294], ['AorrXHXRYoMm', 98], ['AorrXHXRYomM', 98]]
    first index is value and the second index is the number of times it appeared so we have an idea of the chance of
    that genotype appearing

    then disease.disease_posibility returns wheter or not the genotype has the disease
    b is all posibilities and a is the ones that has the disease

    we divide a by b and then multiply with 100 and round with 2 digits after comma
    '''
    result = []
    father = None
    mother = None
    male_child = []
    female_child = []
    for person in Patient.Patients:
        if person.id == father_id:
            father = person

    for person in Patient.Patients:
        if person.id == mother_id:
            mother = person
    if father == None or mother == None:
        return False
    else:
        for gene in Gene.Genes:
            a = gene.gene_in_genos(mother.GenoType, False)
            b = gene.gene_in_genos(father.GenoType, True)
            if len(a) != 0 and len(b) != 0:
                male_child = Patient.Crosser(male_child, [a, b], gene.gonosomal, True)
                female_child = Patient.Crosser(female_child, [a, b], gene.gonosomal, False)
            if gene.letters[0][0] == "Y":
                male_child = Patient.Crosser(male_child, [a, b], gene.gonosomal, True)
        toAdd = []
        print(male_child)
        for disease in Disease.Diseases:
            a = 0
            b = 0
            for geno in male_child:
                if disease.disease_posibility(geno[0], True):
                    a += geno[1]
                b += geno[1]
            res = round((a / b) * 100, 2)
            toAdd.append([disease.name, res])
        result.append(toAdd)
        toAdd = []
        for disease in Disease.Diseases:
            a = 0
            b = 0
            for geno in female_child:
                if disease.disease_posibility(geno[0], False):
                    a += geno[1]
                b += geno[1]
            res = round((a / b) * 100, 2)
            toAdd.append([disease.name, res])
        result.append(toAdd)

        return result


def child_phenotype_posibility(father_id, mother_id):
    #THİS PART İS COPY PASTED FROM child_disease_posibility
    father = None
    mother = None
    male_child = []
    female_child = []
    for person in Patient.Patients:
        if person.id == father_id:
            father = person

    for person in Patient.Patients:
        if person.id == mother_id:
            mother = person
    if father == None or mother == None:
        return False
    else:
        for gene in Gene.Genes:
            a = gene.gene_in_genos(mother.GenoType, False)
            b = gene.gene_in_genos(father.GenoType, True)
            if len(a) != 0 and len(b) != 0:
                male_child = Patient.Crosser(male_child, [a, b], gene.gonosomal, True)
                female_child = Patient.Crosser(female_child, [a, b], gene.gonosomal, False)
            if gene.letters[0][0] == "Y":
                male_child = Patient.Crosser(male_child, [a, b], gene.gonosomal, True)
    for geno in range(len(male_child)):
        toAdd = ""
        for gene in Gene.Genes:
            if gene.gonosomal == False:
                 z = gene.gene_in_genos([male_child[geno][0]])
                 if len(z) > 0:
                      toAdd += gene.geno_to_pheno(z[0])
            elif gene.gonosomal == True:
                z = gene.gene_in_genos([male_child[geno][0]], True)
                if len(z) > 0:
                   toAdd += gene.geno_to_pheno(z[0])
        male_child[geno][0] = toAdd

    for geno in range(len(female_child)):
        toAdd = ""
        for gene in Gene.Genes:
            if gene.gonosomal == False:
                 z = gene.gene_in_genos([female_child[geno][0]])
                 if len(z) > 0:
                    toAdd += gene.geno_to_pheno(z[0])
            elif gene.gonosomal == True:
                z = gene.gene_in_genos([female_child[geno][0]], False)
                if len(z) > 0:
                     toAdd += gene.geno_to_pheno(z[0])
        female_child[geno][0] = toAdd

    female_child = Counter(female_child, mode=True, Count=False)
    male_child = Counter(male_child, mode=True, Count=False)
    a = 0
    b = 0
    fem_list = []
    male_list = []
    for gen in female_child:
        a = gen[1]
        b = 0
        for c in female_child:
            b += c[1]
        fem_list.append([gen[0], round(a*100/b, 2)])
    for gen in male_child:
        a = gen[1]
        b = 0
        for c in male_child:
            b += c[1]
        male_list.append([gen[0], round(a*100/b, 2)])
    print("Female phenotypes and posibilities: {}".format(fem_list))
    print("Male phenotypes and posibilities: {}".format(male_list))
#-----------Front End Funcs---------------------


def main_func():
    #in this function genotype or phenotypes are guessed
    # -----------------------------------First Phase----------------------------------------
    all = []
    fPhase = 0
    sPhase = 0
    try:
        #IN THIS FOR LOOP PATİENTS WİTH GENOTYPES BUT NO PHENOTYPE WİLL HAVE PHENOTYPE ADDED TO THEM
        for person in Patient.Patients:
            toAdd = ""
            x = 0
            g = False
            Otogenes = []
            Gonogenes = []
            if person.GenoType != None and person.PhenoType == None:
                for gene in Gene.Genes:
                    if gene.gonosomal == False:
                        z = gene.gene_in_genos(person.GenoType)
                        if len(z) > 0:
                            toAdd += gene.geno_to_pheno(z[0])
                            c = gene.gene_in_pheno(toAdd)
                            if len(c) > 1:
                                Otogenes.append([c[0], gene])
                                Otogenes.append([c[1], gene])
                            else:
                                Otogenes.append([toAdd, gene])
                    elif gene.gonosomal == True:
                        z = gene.gene_in_genos(person.GenoType, person.Male)
                        if len(z) > 0:
                            toAdd += gene.geno_to_pheno(z[0])
                            c = gene.gene_in_pheno(toAdd)
                            if len(c) > 1:
                                Gonogenes.append([c[0], gene])
                                Gonogenes.append([c[1], gene])
                            else:
                                Gonogenes.append([toAdd, gene])
                all.append([person, Otogenes, Gonogenes])
                person.PhenoType = toAdd
        # IN THIS FOR LOOP PATİENTS WİTH PHENOTYPE BUT NO GENOTYPE WİLL HAVE POSSİBLE GENOTYPES ADDED TO THEM
        for person in Patient.Patients:
            if person.PhenoType != None and person.GenoType == None:
                i = 0
                Pheno = person.PhenoType
                Otogenes = []
                Gonogenes = []
                g = False
                while i < len(Pheno):
                    if (str(Pheno[i]).upper() == "Y" or str(Pheno[i]).upper() == "X") and g == False:
                        x = i + 2
                        g = True
                    else:
                        x = i + 1
                    try:
                        while True:
                            int(Pheno[x])
                            x += 1
                    except:
                        gene = Gene.gene_finder(Pheno[i:x])
                    if gene.gonosomal == False:
                        Otogenes.append([Pheno[i: x], gene])
                    else:
                        Gonogenes.append([Pheno[i: x], gene])
                    g = False
                    i = x
                i = 0
                if person.GenoType == None:
                    while i < len(Otogenes):

                        if i != len(Otogenes) - 1:
                            if Otogenes[i][1] == Otogenes[i + 1][1]:

                                toAdd = Gene.pheno_to_geno(Otogenes[i][1], [Otogenes[i][0], Otogenes[i + 1][0]])
                                person.gene_merger(toAdd)
                                i += 2
                            else:

                                toAdd = Gene.pheno_to_geno(Otogenes[i][1], [Otogenes[i][0]])
                                person.gene_merger(toAdd)
                                i += 1
                        else:

                            toAdd = Gene.pheno_to_geno(Otogenes[i][1], [Otogenes[i][0]])
                            person.gene_merger(toAdd)
                            i += 1
                    i = 0
                    while i < len(Gonogenes):
                        if i < len(Gonogenes) - 1:
                            if Gonogenes[i][1] == Gonogenes[i + 1][1]:
                                a = Gonogenes[i]
                                b = Gonogenes[i + 1]
                                toAdd = Gene.pheno_to_geno(a[1], [Gonogenes[i][0], Gonogenes[i + 1][0]], person.Male)
                                i += 2
                            else:
                                a = Gonogenes[i]
                                toAdd = Gene.pheno_to_geno(a[1], [Gonogenes[i][0]], person.Male)
                                person.gene_merger(toAdd)
                                i += 1
                        else:
                            a = Gonogenes[i]
                            toAdd = Gene.pheno_to_geno(a[1], [Gonogenes[i][0]], person.Male)
                            person.gene_merger(toAdd)
                            i += 1

                all.append([person, Otogenes, Gonogenes])

        print_values()

        # ------------------------------------Second Phase------------------------------------------------
        #IN THIS WHILE LOOP STARTING FROM BIGGEST GENERATION PATIENTS PARENT WILL BE ANALYZED TO REMOVE
        #WRONG GENOTYPES
        bigG = Patient.BigG
        while bigG > 1:
            for person in Patient.Patients:
                if person.Generation == bigG:
                    father = person.Father
                    mother = person.Mother
                    newFather = []
                    newMother = []
                    for i in father.GenoType:
                        newFather.append(i)
                    for i in mother.GenoType:
                        newMother.append(i)
                    childGood = []
                    for gene in Gene.Genes:
                        if person.Male == False or gene.gonosomal == False:
                            childgene = gene.gene_in_genos(person.GenoType, person.Male)
                            mothergene = gene.gene_in_genos(newMother, False)
                            fathergene = gene.gene_in_genos(newFather, True)

                            a = Counter(mothergene)
                            b = Counter(fathergene)
                            c = Counter(childgene)

                            fatherGood = []
                            motherGood = []

                            if len(a) > 0 and len(b) > 0 and len(c) > 0:
                                if gene.gonosomal == False:
                                    for i in range(len(a)):
                                        for x in range(len(b)):
                                            if len(a[i]) != 0 and len(b[x]) != 0:
                                                v = False
                                                for y in c:
                                                    if [a[i][0], b[x][0]] == y or [a[i][0], b[x][1]] == y or [a[i][1],
                                                                                                              b[x][
                                                                                                                  0]] == y or [
                                                        a[i][1], b[x][1]] == y:
                                                        fatherGood.append(b[x])
                                                        motherGood.append(a[i])
                                                        childGood.append(y)
                                                    elif [b[x][0], a[i][0]] == y or [b[x][1], a[i][0]] == y or [b[x][0],
                                                                                                                a[i][
                                                                                                                    1]] == y or [
                                                        b[x][1], a[i][1]] == y:
                                                        fatherGood.append(b[x])
                                                        motherGood.append(a[i])
                                                        childGood.append(y)
                                else:
                                    for i in range(len(a)):
                                        for x in range(len(b)):
                                            if len(a[i]) != 0 and len(b[x]) != 0:
                                                v = False
                                                for y in c:
                                                    if [a[i][0], b[x][0]] == y or [a[i][1], b[x][0]] == y:

                                                        fatherGood.append(b[x])
                                                        motherGood.append(a[i])
                                                        childGood.append(y)
                                                    elif [b[x][0], a[i][0]] == y or [b[x][0], a[i][1]] == y:

                                                        fatherGood.append(b[x])
                                                        motherGood.append(a[i])
                                                        childGood.append(y)

                                motherGood = Counter(motherGood)
                                fatherGood = Counter(fatherGood)
                                i = 0
                                while i < len(fathergene):
                                    v = False
                                    for x in fatherGood:
                                        if fathergene[i] == x:
                                            v = True
                                    if v == False:
                                        fathergene[i] = []
                                        i += 1
                                    else:
                                        i += 1
                                i = 0
                                while i < len(mothergene):
                                    v = False
                                    for x in motherGood:
                                        if mothergene[i] == x:
                                            v = True
                                    if v == False:
                                        mothergene[i] = []
                                        i += 1
                                    else:
                                        i += 1
                                i = 0
                                while i < len(newFather):
                                    if fathergene[i] == []:
                                        newFather.pop(i)
                                        fathergene.pop(i)
                                    else:
                                        i += 1
                                i = 0
                                while i < len(newMother):
                                    if mothergene[i] == []:
                                        newMother.pop(i)
                                        mothergene.pop(i)
                                    else:
                                        i += 1
                        else:
                            childgene = gene.gene_in_genos(person.GenoType, True)
                            fathergene = gene.gene_in_genos(newFather, True)
                            mothergene = gene.gene_in_genos(newMother, False)

                            a = Counter(mothergene)
                            b = Counter(fathergene)
                            c = Counter(childgene)

                            if len(a) == 0:

                                if b != c:
                                    Patient.MutatedPatients.append(person)
                            else:
                                for i in range(len(a)):

                                    if a[i][0] != c[0][0] and a[i][1] != c[0][0]:
                                        a[i] = []
                                for i in range(len(mothergene)):
                                    v = False
                                    for b in a:
                                        if mothergene[i] == b:
                                            v = True
                                    if v != True:
                                        mothergene[i] = []
                                i = 0
                                while i < len(mothergene):
                                    if mothergene[i] == []:
                                        newMother.pop(i)
                                        mothergene.pop(i)
                                    else:
                                        i += 1

                    if len(newMother) > 0:
                        mother.GenoType = newMother
                    elif len(childgene) > 0:

                        Patient.MutatedPatients.append(person)
                    if len(newFather) > 0:
                        father.GenoType = newFather
                    elif len(childgene) > 0:
                        Patient.MutatedPatients.append(person)
            bigG -= 1
        Patient.MutatedPatients = Counter(Patient.MutatedPatients)

        # IN THIS WHILE LOOP STARTING FROM BIGGEST GENERATION PATIENTS CHILDREN WILL BE ANALYZED TO REMOVE
        # WRONG GENOTYPES
        bigG = Patient.BigG
        while bigG > 1:
            for person in Patient.Patients:
                v = True
                for i in Patient.MutatedPatients:
                    if person == i:
                        v = False
                if v and person.Generation == bigG:
                    childGood = []
                    if person.Male == False:
                        for gene in Gene.Genes:
                            fathergene = person.Father.GenoType
                            fathergene = Gene.gene_in_genos(gene, fathergene)
                            fathergene = Counter(fathergene)

                            mothergene = person.Mother.GenoType
                            mothergene = Gene.gene_in_genos(gene, mothergene)
                            mothergene = Counter(mothergene)

                            Child = Gene.gene_in_genos(gene, person.GenoType)
                            for a in mothergene:
                                for b in fathergene:
                                    for c in Child:
                                        if [a[0], b[0]] == c or [a[1], b[0]] == c or [a[0], b[1]] == c or [a[1],b[1]] == c:

                                            childGood.append(c)
                                        elif [b[0], a[0]] == c or [b[0], a[1]] == c or [b[1], a[0]] == c or [b[1], a[1]] == c:

                                            childGood.append(c)
                    else:
                        for gene in Gene.Genes:
                            if gene.gonosomal == False:

                                fathergene = person.Father.GenoType
                                fathergene = Gene.gene_in_genos(gene, fathergene)
                                fathergene = Counter(fathergene)

                                mothergene = person.Mother.GenoType
                                mothergene = Gene.gene_in_genos(gene, mothergene)
                                mothergene = Counter(mothergene)

                                Child = Gene.gene_in_genos(gene, person.GenoType)
                                Child = Counter(Child)

                                if len(Child) > 0 and len(fathergene) > 0 and len(mothergene) > 0:
                                    for a in mothergene:
                                        for b in fathergene:
                                            for c in Child:
                                                if [a[0], b[0]] == c or [a[1], b[0]] == c or [a[0], b[1]] == c or [a[1],
                                                                                                                   b[
                                                                                                                       1]] == c:
                                                    childGood.append(c)
                                                elif [b[0], a[0]] == c or [b[1], a[0]] == c or [b[0], a[1]] == c or [
                                                    b[1], a[1]] == c:
                                                    childGood.append(c)

                    newChild = person.GenoType
                    childGood = Counter(childGood)
                    for good in childGood:
                        goods = []
                        gene = Gene.gene_finder(good[0])
                        childgene = Gene.gene_in_genos(gene, newChild)
                        for i in range(len(childgene)):
                            if childgene[i] == good:
                                goods.append(i)
                        c = []
                        for a in range(len(newChild)):
                            p = True
                            for i in goods:
                                if a == i:
                                    p = False
                            if p == False:
                                c.append(newChild[a])
                        newChild = c
                    if len(newChild) > 0:
                        person.GenoType = newChild
            bigG -= 1
    except AttributeError:
        errors.append(Lang[0][n].format(person.Fullname, person.id))


def child_disease(father_id, mother_id):
    '''
    FE FUNC
    Takes id of father and mother then sends them to child_disease_posibility
    and writes the results to a txt file inside reports
    '''
    fatherName = None
    motherName = None
    result = child_disease_posibility(int(father_id), int(mother_id))
    if result == False:
        return False
    for person in Patient.Patients:
        if person.id == father_id:
            fatherName = person.Fullname
        if person.id == mother_id:
            motherName = person.Fullname

    report = open(getPath() + "\{}".format(Lang[6][n])+"\{}".format(Lang[8][n].format(motherName, fatherName)) +".txt",
                  mode="w",
                  encoding="utf-8")
    report.write(Lang[9][n])
    report.write("\n")
    for dis in result[0]:
        report.write("{}: %{}\n".format(dis[0], dis[1]))
    report.write("\n{}".format(Lang[10][n]))
    report.write("\n")
    for dis in result[1]:
        report.write("{}: %{}\n".format(dis[0], dis[1]))
    return None


def decoder(mode=0, change=True):
    #MODE 0 CYPHERS ALL NAMES AND REMOVES THEM FROM THE EXCEL
    #MODE 1 RETURNS A LİST OF NAMES
    #MODE 2 REWRİTES ALL NAMES TO EXCEL FİLE
    global n
    #8 bit codes of letters and numbers
    binary = [
        ["A", "01000001"], ["B", "01000010"], ["C", "01000011"], ["D", "01000100"],
        ["E", "01000101"], ["F", "01000110"], ["G", "01000111"], ["H", "01001000"],
        ["I", "01001001"], ["J", "01001010"], ["K", "01001011"], ["L", "01001100"],
        ["M", "01001101"], ["N", "01001110"], ["O", "01001111"], ["P", "01010000"],
        ["Q", "01010001"], ["R", "01010010"], ["S", "01010011"], ["T", "01010100"],
        ["U", "01010101"], ["V", "01010110"], ["W", "01010111"], ["X", "01011000"],
        ["Y", "01011001"], ["Z", "01011010"], ["a", "01100001"], ["b", "01100010"],
        ["c", "01100011"], ["d", "01100100"], ["e", "01100101"], ["f", "01100110"],
        ["g", "01100111"], ["h", "01101000"], ["i", "01101001"], ["j", "01101010"],
        ["k", "01101011"], ["l", "01101100"], ["m", "01101101"], ["n", "01101110"],
        ["o", "01101111"], ["p", "01110000"], ["q", "01110001"], ["r", "01110010"],
        ["s", "01110011"], ["t", "01110100"], ["u", "01110101"], ["v", "01110110"],
        ["w", "01110111"], ["x", "01111000"], ["y", "01111001"], ["z", "01111010"],
        ["ö", "00100001"], ["Ö", "00100010"], ["İ", "00100011"], ["ı", "00100100"],
        ["ü", "00100101"], ["Ü", "00100110"], ["ğ", "00100111"], ["Ğ", "00101000"],
        ["ş", "00101001"], ["Ş", "00101010"], ["ç", "00101011"], ["Ç", "00101100"],
        ["1", "00110001"], ["2", "00110010"], ["3", "00110011"], ["4", "00110100"],
        ["5", "00110101"], ["6", "00110110"], ["7", "00110111"], ["8", "00111000"],
        ["9", "00111001"], ["0", "00110000"]
    ]
    if mode == 0:

        excel = openpyxl.load_workbook("1.xlsx")
        ws = excel.worksheets[0]

        data = open(getPath() + "\{}\data1.txt".format(Lang[5][n]), "w")
        toWrite = ""
        y = 2
        while ws.cell(y, 1).value != None:
            name = ws.cell(y, 1).value
            for i in name:
                for b in binary:
                    if i == b[0]:
                        toWrite = toWrite + b[1]
                    elif i == " ":
                        toWrite = toWrite + " "
            ws.cell(y, 1).value = "Hidden"
            y += 1
            toWrite += ","
        toWrite += "."
        data.write(toWrite)
        data.close()
        excel.save("1.xlsx")
        if change:

            opt = open(getPath() +"\{}\options.txt".format(Lang[5][n]), "w")
            opt.write("Privacy: {}\nLang: {}".format(dataPrivacy, n))

    elif mode == 1:

        names = []
        data = open(getPath() + "\{}\data1.txt".format(Lang[5][n]), "r")
        z = 0
        end = False
        toWrite = ""
        while end is False:
            a = data.read(1)
            if a == ",":
                z += 1
                names.append(toWrite)
                toWrite = ""
            elif a == ".":
                end = True
            elif a == " ":
                toWrite += " "
            else:
                data.seek(z)
                letter = data.read(8)
                for i in binary:
                    if letter == i[1]:
                        toWrite += i[0]
                z += 8
            data.seek(z)
        return names

    elif mode == 2:

        names = decoder(1)
        excel = openpyxl.load_workbook("1.xlsx")
        ws = excel.worksheets[0]
        y = 2
        for name in names:
            ws.cell(y, 1).value = name
            y += 1
        excel.save("1.xlsx")
        if change:

            opt = open(getPath() + "\{}\options.txt".format(Lang[5][n]), "w")
            opt.write("Privacy: {}\nLang: {}".format(dataPrivacy, n))


def new_excel():
    #TURNS EXCEL DATA İNTO CLASS İNSTANCES OF PATİENT, DİSEASE AND GENE
    Patient.Patients = []
    Patient.MutatedPatients = []
    Disease.symptoms = []
    Disease.Diseases = []
    Gene.Genes = []
    if dataPrivacy == 0:
        decoded = False
    else:
        decoded = True
    if decoded:
        names = decoder(1)

    if True:
        excel = openpyxl.load_workbook("1.xlsx")
        ws = excel.worksheets[0]
        y = 2
        while ws.cell(y, 1).value != None:
            #PATİENTS ARE DEFİNED HERE
            if ws.cell(y, 4).value == None:
                geno = None
            else:
                geno = [ws.cell(y, 4).value]
            if ws.cell(y, 7).value == None:
                sex = False
            elif str(ws.cell(y, 7).value).upper() == "E":
                sex = True
            else:
                sex = False
            if decoded is False:
                Patient(ws.cell(y, 1).value, None, None, geno, ws.cell(y, 5).value, ws.cell(y, 6).value, sex)
            else:
                Patient(names[y-2], None, None, geno, ws.cell(y, 5).value, ws.cell(y, 6).value, sex)
            y += 1
        y = 2
        while ws.cell(y, 1).value != None:
            if ws.cell(y, 2).value != None:
                perid = ws.cell(y, 6).value
                for person in Patient.Patients:
                    if person.id == perid:
                        per = person

                b = ws.cell(y, 2).value
                a = ws.cell(y, 3).value
                for person in Patient.Patients:
                    if person.id == b:
                        per.Father = person
                        person.Children.append(per)
                    elif person.id == a:
                        per.Mother = person
                        person.Children.append(per)

            y += 1
        try:
            for person in Patient.Patients:
                if (person.Mother != None and person.Father == None) or (
                        person.Father != None and person.Mother == None):
                    raise ZeroDivisionError
        except ZeroDivisionError:
            errors.append(
                Lang[1][n].format(person.id, person.Fullname))
        for person in Patient.Patients:
            if person.Mother == None and person.Father == None:
                person.Generation = 1
        a = False

        while a == False and len(errors) == 0:
            #BigG variable and generation of all patients are defined here
            for person in Patient.Patients:
                if person.Mother != None and person.Father != None:
                    if person.Mother.Generation != None and person.Father.Generation != None:
                        if person.Father.Generation > person.Mother.Generation:
                            person.Generation = person.Father.Generation + 1
                        else:
                            person.Generation = person.Mother.Generation + 1
                        if person.Generation > Patient.BigG:
                            Patient.BigG = person.Generation
            a = True
            for person in Patient.Patients:
                if person.Generation == None:
                    a = False


        ws = excel.worksheets[1]
        y = 3
        while ws.cell(y, 1).value != None:
            #GENES ARE DEFİNED HERE
            dominance = []
            x = str(ws.cell(y, 3).value)
            for i in x:
                dominance.append(int(i))
            letters = (str(ws.cell(y, 2).value).replace(" ", "")).split(",")
            a = False
            if letters[0][0] == "X" or letters[0][0] == "Y":
                a = True
            Gene(ws.cell(y, 1).value, dominance, letters, a)
            y += 1
        y = 3
        ws = excel.worksheets[2]

        while ws.cell(y, 1).value != None:
            #DİSEASES ARE DEFİNED HERE
            Disease(ws.cell(y, 1).value, ws.cell(y, 2).value)
            y += 1


def show_person(id, mode=0):
    #WRİTES TXT FİLE ABOUT THE PERSON WİTH İD, REPORT İS LOCATED İNSİDE REPORTS FİLE
    if mode == 0:
        target = None
        for person in Patient.Patients:
            if person.id == id:
                target = person
        if target == None:
            return False
        else:
            if target.Male == True:
                x = "male"
            else:
                x = "female"
            report = open(getPath()+"\{}".format(Lang[6][n])+"\{}".format(target.Fullname) + " " + Lang[2][n] + ".txt", mode="w", encoding="utf-8")
            if target.Mother != None and target.Father != None:
                report.write(Lang[3][n].format(target.Fullname, target.id,x,target.Mother.Fullname,
                                               target.Mother.id, target.Father.Fullname, target.Father.id,
                                               target.PhenoType))
            else:
                report.write(Lang[4][n].format(target.Fullname, target.id, x,
                                               target.PhenoType))
            i = 0
            report.write("\n")
            while i < len(target.GenoType):

                if i % 2 != 0:
                    report.write("\t"+target.GenoType[i]+"\n")
                    i += 1
                else:
                    report.write(target.GenoType[i])
                    i += 1
            report.close()
            return None
    elif mode == 1:
        for target in Patient.Patients:
            if target.Male == True:
                x = "male"
            else:
                x = "female"
            report = open(getPath() + "\{}".format(Lang[6][n]) + "\{}".format(target.Fullname) + " " + Lang[2][n] + ".txt", mode="w",
                          encoding="utf-8")
            if target.Mother != None and target.Father != None:
                report.write(Lang[3][n].format(target.Fullname, target.id, x, target.Mother.Fullname,
                                               target.Mother.id, target.Father.Fullname, target.Father.id,
                                               target.PhenoType))
            else:
                report.write(Lang[4][n].format(target.Fullname, target.id, x,
                                               target.PhenoType))
            i = 0
            report.write("\n")
            while i < len(target.GenoType):

                if i % 2 != 0:
                    report.write("\t" + target.GenoType[i] + "\n")
                    i += 1
                else:
                    report.write(target.GenoType[i])
                    i += 1
            report.close()
        return None


def LangChange(newN):
    #renames the file names system and reports
    global n
    a = n
    rename(Lang[5][a], Lang[5][newN])
    rename(Lang[6][a], Lang[6][newN])
    n = newN


def show_disease(id=None, mode=0):
    #FE func, takes an id and writes all diseases of person into a txt file found inside reports/raporlar
    if mode == 0:
        target = None
        for person in Patient.Patients:
            if person.id == id:
                target = person
        if target == None:
            return False

        report = open(getPath() + "\{}".format(Lang[6][n]) + "\{}".format(target.Fullname) + " " + Lang[7][n] + ".txt",
                      mode="w",
                      encoding="utf-8")
        report.write(target.Fullname)
        for dis in disease_posibility_person(id):
            report.write("\n")
            report.write("{}: %{}".format(dis[0], dis[1]))
        report.close()
    elif mode == 1:
        for target in Patient.Patients:
            report = open(
                getPath() + "\{}".format(Lang[6][n]) + "\{}".format(target.Fullname) + " " + Lang[7][n] + ".txt",
                mode="w",
                encoding="utf-8")
            report.write(target.Fullname)
            for dis in disease_posibility_person(target.id):
                report.write("\n")
                report.write("{}: %{}".format(dis[0], dis[1]))
            report.close()
    return None


def givePersonInfo(id):
    #FE FUNC, takes an id and returns information of the person
    excel = openpyxl.load_workbook("1.xlsx")
    ws = excel.worksheets[0]
    y = 2
    while ws.cell(y, 1).value != None:
        if int(ws.cell(y, 6).value) == id:
            geno = ws.cell(y, 4).value
            pheno = ws.cell(y, 5).value
            if geno == None:
                geno = ""
            if pheno == None:
                pheno = ""
            if dataPrivacy == 1:
                for person in Patient.Patients:
                    if person.id == int(id):
                        name = person.Fullname
            else:
                name = ws.cell(y, 1).value
            return [name, geno, pheno, str(ws.cell(y, 6).value), ws.cell(y, 7).value]
        y += 1
    return False


def AppylPerson(name, Geno, Pheno, id, sex):
    #FE FUNC, takes input from front end and rewrites the excel with the changed values

    if dataPrivacy == 1:
        try:
            decoder(2, change=False)
        except:
            return Lang[11][n]
    excel = openpyxl.load_workbook("1.xlsx")
    ws = excel.worksheets[0]
    id = int(id)
    if str(sex).upper() != "K" and str(sex).upper() != "E":
        return False
    sex = str(sex).upper()
    y = 2
    while ws.cell(y, 1).value != None:
        if int(ws.cell(y, 6).value) == id:
            ws.cell(y, 1).value = name
            if Geno == "":
                ws.cell(y, 4).value = None
            else:
                ws.cell(y, 4).value = Geno
            if Pheno == "":
                ws.cell(y, 5).value = None
            else:
                ws.cell(y, 5).value = Pheno
            ws.cell(y, 7).value = sex
        y += 1
    try:
        excel.save("1.xlsx")
        if dataPrivacy == 1:
            decoder(0, change=False)
        return True
    except:
        return Lang[11][n]


#---------------Code Trash----------------


def ChangeGeneLetter(geneName, newLetters):
    #FE Func
    '''
    Changes a genes letters for example: ["A", "B"] to ["C", "D"]
    then changes the letters in every place they are used from peoples pheno or genotypes to disease phenotypes
    INPUT: geneName = "BloodType"
            newLetters = ["c", "d", "O"]

    OUTPUT: if done True else False
    '''
    target = None
    for gene in Gene.Genes:
        if gene.name == geneName:

            target = gene
    excel = openpyxl.load_workbook("1.xlsx")
    ws = excel.worksheets[1]
    y = 3
    while ws.cell(y, 1).value != None:
        if ws.cell(y, 1).value == geneName:
            ws.cell(y, 2).value = newLetters

        y += 1
    newLetters = str(newLetters).replace(" ", "")
    newLetters = str(newLetters).split(",")

    if len(target.letters) != len(newLetters):
        return False
    for i in newLetters:
        if len(i) == 0:
            return False
    oldLetters = target.letters
    y = 2

    ws = excel.worksheets[0]
    y = 2
    while ws.cell(y, 1).value != None:

        geno = ws.cell(y, 4).value
        pheno = ws.cell(y, 5).value
        for i in range(len(oldLetters)):
            a = 0
            b = len(oldLetters[i])
            if geno != None:
                while b <= len(geno):
                    print(a, b)
                    print(geno[a: b])
                    if geno[a:b] == oldLetters[i]:
                        if geno[a - 1] != "Y" and geno[a - 1] != "X":
                            try:
                                print(geno[b])
                                int(geno[b])
                                a = b
                                b += len(oldLetters[i])
                            except:

                                jump = len(geno[:a]) + len(newLetters[i])
                                geno = geno[:a] + newLetters[i] + geno[b:]
                                a = jump
                                b = jump + len(oldLetters[i])
                        else:
                            a += 1
                            b += 1


                    else:
                        a += 1
                        b += 1
                ws.cell(y, 4).value = geno
            a = 0
            b = len(oldLetters[i])
            if pheno != None:
                while b <= len(pheno):
                    if pheno[a:b] == oldLetters[i]:
                        if pheno[a-1] != "Y" and pheno[a-1] != "X":
                            try:
                                int(pheno[b])
                                a = b
                                b += len(oldLetters[i])
                            except:
                                jump = len(pheno[:a]) + len(newLetters[i])
                                pheno = pheno[:a] + newLetters[i] + pheno[b:]
                                a = jump
                                b =jump + len(oldLetters[i])
                        else:
                            a += 1
                            b += 1
                    else:
                        a += 1
                        b += 1
                ws.cell(y, 5).value = pheno

        y += 1
    ws = excel.worksheets[2]
    y = 3

    while ws.cell(y, 1).value != None:
        pheno = ws.cell(y, 2).value
        for i in range(len(oldLetters)):
            a = 0
            b = len(oldLetters[i])

            while b <= len(pheno):

                if pheno[a:b] == oldLetters[i]:
                    if pheno[a - 1] != "Y" and pheno[a - 1] != "X":

                        try:
                            int(pheno[b + 1])
                        except:
                            jump = len(pheno[:a]) + len(newLetters[i])
                            pheno = pheno[:a] + newLetters[i] + pheno[b:]
                            a = jump
                            b = jump + len(oldLetters[i])
                    else:
                        a+=1
                        b+=1
                else:
                    a += 1
                    b += 1
        ws.cell(y, 2).value = pheno
        y += 1
    try:
        excel.save("1.xlsx")
    except:
        return Lang[11][n]
    return True


def OldLetters(geneName):
    #returns the current letters from the gene
    #called from ChangeGeneLetter
    x = ""
    for gene in Gene.Genes:
        if gene.name == geneName:
            x = gene.letters[0]
            for i in range(len(gene.letters)):
                if i != 0:
                    x += ","+gene.letters[i]
    return x


def allGenes():
    #returns all genes
    g = []
    for gene in Gene.Genes:
        g.append(gene.name)
    return g

if __name__ == '__main__':
    new_excel()
    main_func()
    print_values()
'''''
            x = 0
            gene = Gene.gene_finder(Pheno[i:i+x+1])
            while gene == None:
                x += 1
                gene = Gene.gene_finder(Pheno[i:i + x + 1])
            print(person.Fullname)
            print(gene.name)
            print(person.PhenoType)
            nextgen = Gene.gene_finder(i + x + 1)

            if i < len(Pheno) - 1:
                if Gene.gene_finder(Pheno[i+1]) == gene:
                    genes = gene.geno_type_finder(Pheno[i:i + 2])
                    person.gene_merger(genes)
                    i += 2
                else:
                    print(Pheno[i])
                    genes = gene.geno_type_finder(Pheno[i])
                    person.gene_merger(genes)
                    i += 1
            else:
                genes = gene.geno_type_finder(Pheno[i])
                person.gene_merger(genes)
                i += 1
            '''''
'''
        numerical = ""
        if len(gene) > 1:
            return [gene]
        posibilities = []
        for i in gene:
            for a in range(0, len(self.letters)):
                if i == self.letters[a]:
                    numerical = numerical + str(self.dominance[a])
        posibilities.append("{}{}".format(gene, gene))
        for a in range(len(self.dominance)):
            if self.dominance[a] < int(numerical):
                posibilities.append("{}{}".format(gene, self.letters[a]))
        return posibilities
        '''
'''
    gene1 = Gene("Blood Type", [1, 1, 0], ["A", "B", "O"])
    gene2 = Gene("RH", [1, 0], ["R", "r"])
    gene3 = Gene("Gonosomal1", [1, 0], ["YN", "Yn"], True)
    gene4 = Gene("Gonosomal2", [1,0], ["XA", "Xa"], True)
    fatma = Patient("Fatma Soy", None, None, None, "ARXA")
    murat = Patient("Murat Soy", None, None, ["BBRrXaYN"], None)
    burak = Patient("Burak Soy", murat, fatma, None, "BRXAYN")
    melis = Patient("Melis Soy", None, None, None, "OrXa")
    firat = Patient("Fırat Soy", burak, melis, None, "ORXAYn")
    sabri = Patient("Sabri Soy", burak, melis, None, "BRXaYN")
    kubra = Patient("Kübra Soy", burak, melis, ["OOrrXaXN"], None)
    sila = Patient("Sıla Balcı", murat, fatma, None, "ABRXa")
    semih = Patient("Semih Balcı", None, None, None, None)
    senem = Patient("Senem Balcı", semih, sila, None, None)
    kamil = Patient("Kamil Balcı", semih, sila, None, "BRXaYn")
    '''
'''
gen1 = Gene("Blood Type", [1, 1, 0], ["A", "B", "o"])
    gen2 = Gene("RH", [1, 0], ["R", "r"])
    gen3 = Gene("Hemophili", [1, 0], ["H", "h"])
    gen4 = Gene("Color Blindness", [1, 0], ["XR", "Xr"], True)
    gen5 = Gene("Obez", [1, 0], ["YO", "Yo"], True)
    gen6 = Gene("brr", [2, 1, 0], ["r3", "r2", "r1"])
kadın1 = Patient("Kadın1", None, None, None, "ArHXR")
    adam1 = Patient("Adam1", None, None, None, "BrHXrYo", True)
    adam3 = Patient("Adam3", adam1, kadın1, None, "OhrXRYo", True)
    kadın2 = Patient("Kadın2", adam1, kadın1, None, "ArHXR")
    adam4 = Patient("Adam4", adam1, kadın1, None, "ABrHXRYo", True)
    adam5 = Patient("adam5", None, None, None, "ARHXrYO", True)
    kadın3 = Patient("Kadın3", None, None, None, "ARHXR")
    kadın4 = Patient("Kadın4", adam5, kadın3, None, "ORHXR")
    adam6 = Patient("Adam6", adam5, kadın3, None, "ARHXrYO", True)
    kadın5 = Patient("Kadın5", adam6, kadın2, None, "OrHXr")
    adam7 = Patient("Adam7", adam6, kadın2, None, "ARhXRYO", True)
    adam8 = Patient("Adam8", None, None, None, "ABRHXrYO", True)
    kadın6 = Patient("Kadın6", None, None, None, "BrhXr")
    kadın7 = Patient("Kadın7", adam8, kadın6, None, "ABRHXR")
    adam9 = Patient("Adam9", adam8, kadın6, None, "ARHXrYO", True)
    adam10 = Patient("Adam10", None, None, ["AOrrHhXrYo"], None, True)
'''
'''
bigG = 1
        while bigG <= Patient.BigG:
            for person in Patient.Patients:
                if person.Generation == bigG:
                    for child in person.Children:
                        a = True
                        for i in Patient.MutatedPatients:
                            if i == child:
                                a = False
                        if a == True:
                            mother = None
                            father = None
                            if person.Male == False:
                                mother = person
                                father = child.Father
                            else:
                                father = person
                                mother = child.Mother
'''
'''
Geno = str(person.GenoType[0])
i = 0
while i < len(Geno):
    if (Geno[i].upper() == "X" or Geno[i].upper() == "Y") and g == False:
        x = i + 2
        g = True
    else:
        x = i + 1
    try:
        while True:
            int(Geno[x])
            x += 1
    except:
        gene = Gene.gene_finder(Geno[i:x])
    if gene.gonosomal == False:
        Otogenes.append([Geno[i: x], gene])
    else:
        Gonogenes.append([Geno[i: x], gene])
    g = False
    i = x
i = 0
while i < len(Otogenes):
    toAdd = Otogenes[i][1].pheno_to_geno([Otogenes[i][0], Otogenes[i + 1][0]])
    if person.PhenoType == None:
        person.PhenoType = toAdd
    else:
        person.PhenoType += toAdd
    i += 2
i = 0
while i < len(Gonogenes):
    if i < len(Gonogenes) - 1:
        if Gonogenes[i][1] == Gonogenes[i + 1][1]:
            toAdd = Gonogenes[i][1].geno_to_pheno([Gonogenes[i][0], Gonogenes[i + 1][0]])
            person.PhenoType += toAdd
            i += 2
        else:
            person.PhenoType += Gonogenes[i][0]
            i += 1
    else:
        a = Gonogenes[i][0]
        person.PhenoType += Gonogenes[i][0]
        i += 1
'''
