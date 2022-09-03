import Genotype
import random
import openpyxl

Genotype.new_excel()
Genotype.main_func()
x = 302
def ManAdder(human, sex, motherid, fatherid):
    global x
    excel = openpyxl.load_workbook("1.xlsx")
    ws = excel.worksheets[0]
    y = 1
    while ws.cell(y, 1).value != None:
        y += 1
    if sex == 0:
        ws.cell(y, 1).value = "adam{}".format(x)
        ws.cell(y, 2).value = fatherid
        ws.cell(y, 3).value = motherid
        ws.cell(y, 4).value = None
        ws.cell(y, 5).value =  human
        ws.cell(y, 6).value = x
        ws.cell(y, 7).value = "E"
    if sex == 1:
        ws.cell(y, 1).value = "kadın{}".format(x)
        ws.cell(y, 2).value = fatherid
        ws.cell(y, 3).value = motherid
        ws.cell(y, 4).value = None
        ws.cell(y, 5).value = human
        ws.cell(y, 6).value = x
        ws.cell(y, 7).value = "K"
    x += 1
    excel.save("1.xlsx")
    Genotype.new_excel()
    Genotype.main_func()


while len(Genotype.Patient.Patients) < 1000:
    while True:
        id2 = random.randint(0, len(Genotype.Patient.Patients) - 1)
        if Genotype.Patient.Patients[id2].Male == True:
            father = Genotype.Patient.Patients[id2]
            break
    while True:
        id1 = random.randint(0, len(Genotype.Patient.Patients) - 1)
        if Genotype.Patient.Patients[id1].Male == False:
            mother = Genotype.Patient.Patients[id1]
            break
    sex = random.randint(0, 1)
    male_child = []
    female_child = []
    for gene in Genotype.Gene.Genes:
        a = gene.gene_in_genos(mother.GenoType, False)
        b = gene.gene_in_genos(father.GenoType, True)
        if len(a) != 0 and len(b) != 0:
            male_child = Genotype.Patient.Crosser(male_child, [a, b], gene.gonosomal, True)
            female_child = Genotype.Patient.Crosser(female_child, [a, b], gene.gonosomal, False)
        if gene.letters[0][0] == "Y":
            male_child = Genotype.Patient.Crosser(male_child, [a, b], gene.gonosomal, True)
    toAdd = ""
    if sex == 0:
        geno = random.randint(0, len(male_child) - 1)
        for gene in Genotype.Gene.Genes:
            if gene.gonosomal == False:
                z = gene.gene_in_genos([male_child[geno][0]])
                if len(z) > 0:
                    toAdd += gene.geno_to_pheno(z[0])

            elif gene.gonosomal == True:
                z = gene.gene_in_genos([male_child[geno][0]], True)
                if len(z) > 0:
                    toAdd += gene.geno_to_pheno(z[0])

        ManAdder(toAdd, sex, mother.id, father.id)
    if sex == 1:
        geno = random.randint(0, len(female_child) - 1)
        for gene in Genotype.Gene.Genes:
            if gene.gonosomal == False:
                z = gene.gene_in_genos([female_child[geno][0]])
                if len(z) > 0:
                    toAdd += gene.geno_to_pheno(z[0])
            elif gene.gonosomal == True:
                z = gene.gene_in_genos([female_child[geno][0]], False)
                if len(z) > 0:
                    toAdd += gene.geno_to_pheno(z[0])

        ManAdder(toAdd, sex, mother.id, father.id)
    print("sıra: {}".format(len(Genotype.Patient.Patients)))
